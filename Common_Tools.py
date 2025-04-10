import pandas as pd
import numpy as np
import inspect, re
import openpyxl 
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.styles import Alignment, PatternFill, Border, Side
import csv
import xlsxwriter
import magic 
import random 
from random import randint
from random import uniform
from scipy import stats
import shap
from scipy import stats
import os
import joblib as joblib
from joblib import dump, load
import json
import tkinter as tk
from tkinter import *
import time
import psutil
from PIL import ImageTk, Image
import statistics
from IPython.display import display
#np.random.seed(1000)
rstate = 12

def sanitize_filename(filename):
    """Remove or replace invalid characters from filenames."""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        filename = filename.replace(char, '_')
    return filename

# Python code to remove whitespace
def remove(string):
    return string.replace(" ", "")

def apply_coding(value, bins, code):
    for i, b in enumerate(bins):
        if value in bins[i]:
            return code[i]
        
def varname(p):
    for line in inspect.getframeinfo(inspect.currentframe().f_back)[3]:
        m = re.search(r'\bvarname\s*\(\s*([A-Za-z_][A-Za-z0-9_]*)\s*\)', line)
        if m:
            return m.group(1)

# Function to get machine specs
def get_machine_specs():
    cpu_info = psutil.cpu_percent()
    mem_info = psutil.virtual_memory().total / (1024 ** 3)  # Convert bytes to GB
    return cpu_info, mem_info

def dict_to_excel(data, file_path=""):
    # Convert dictionary to DataFrame
    df = pd.DataFrame(data).T

    # Flatten lists in the DataFrame (if any)
    for column in df.columns:
        df[column] = df[column].apply(lambda x: str(x) if isinstance(x, list) else x)

    # Write DataFrame to Excel
    df.to_excel(file_path)
    
    return df

def wrap_text_excel(file_path):
    # Load the workbook and the sheet
    wb = (file_path)
    ws = wb.active

    # Set text wrapping for all cells and adjust column width based on the longest cell in each column
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter

        for cell in col:
            # Set wrap_text to True for all cells
            cell.alignment = Alignment(wrap_text=True)
            
            # Check the length of the cell value (if it's not None)
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

        # Adjust the column width based on the longest value, considering an approximation factor
        adjusted_width = (max_length + 2) * 1.1  # Add padding and adjust width factor
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save the workbook
    wb.save(file_path)

def expand_cell_excel(file_path):
    # Load the workbook and the sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Adjust the width of the columns based on the max length in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(file_path)

def grid_excel(file_path):
    # Load the workbook and the sheet
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define border style
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))
    
    # Apply borders to all cells to create a grid
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            
    # Save the workbook
    wb.save(file_path)

# Function to fill cells with color if their values are in the target list
def fill_cells_with_color(file_path, target_values, color):
    # Load the workbook and select the specified sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Define the fill pattern with the specified color
    fill_pattern = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Iterate over all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # If the cell value is in the target list, apply the fill pattern
            if cell.value in target_values:
                cell.fill = fill_pattern

    # Save the modified workbook
    workbook.save(file_path)

def parse_exp_csv(df, filename, project_name, unique_value_threshold=10): # Define the threshold for considering discrete numeric columns as categorical
    print(filename[-4:])
    # upload index file
    pathname = project_name + '/' + filename[:-4] + "_index.csv"
    df_exp = upload_data(pathname)
    
    # inputs and outputs for a row
    # return index
    input_cols = list(df_exp.columns[(df_exp == 1).all()])
    outputs_cols = list(df_exp.columns[(df_exp == 2).all()])[0]
    
    input_df = df.loc[:, input_cols]

    # Select columns of object data type (categorical columns)
    categorical_cols = input_df.select_dtypes('object').columns.tolist()

    # Identify discrete numeric columns as categorical if they have fewer unique values than the threshold
    discrete_numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if len(input_df[col].unique()) < unique_value_threshold]
    categorical_cols.extend(discrete_numeric_cols)

    # Exclude the identified discrete numeric columns from numeric_cols
    numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if col not in discrete_numeric_cols]
    
    return input_cols, outputs_cols, categorical_cols, numeric_cols

def parse_exp(df, filename, project_name, unique_value_threshold=10): # Define the threshold for considering discrete numeric columns as categorical
    file_type = filename[-4:]
    print("File Type:", file_type)
    # upload index file
    if file_type == "xlsx":
        pathname = project_name + '/' + filename[:-5] + "_index.xlsx"
    else:
        pathname = project_name + '/' + filename[:-4] + "_index.csv"
    df_exp = upload_data(pathname)
    
    # inputs and outputs for a row
    # return index
    input_cols = list(df_exp.columns[(df_exp == 1).all()])
    outputs_cols = list(df_exp.columns[(df_exp == 2).all()])[0]
    
    input_df = df.loc[:, input_cols]

    # Select columns of object data type (categorical columns)
    categorical_cols = input_df.select_dtypes('object').columns.tolist()

    # Identify discrete numeric columns as categorical if they have fewer unique values than the threshold
    discrete_numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if len(input_df[col].unique()) < unique_value_threshold]
    categorical_cols.extend(discrete_numeric_cols)

    # Exclude the identified discrete numeric columns from numeric_cols
    numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if col not in discrete_numeric_cols]
    
    return input_cols, outputs_cols, categorical_cols, numeric_cols

def parse_exp_multi_outcomes(df, filename, project_name, unique_value_threshold=10): # Define the threshold for considering discrete numeric columns as categorical
    file_type = filename[-4:]
    print("File Type:", file_type)
    # upload index file
    if file_type == "xlsx":
        pathname = project_name + '/' + filename[:-5] + "_index.xlsx"
    else:
        pathname = project_name + '/' + filename[:-4] + "_index.csv"
    df_exp = upload_data(pathname)
    
    # inputs and outputs for a row
    # return index
    input_cols = list(df_exp.columns[(df_exp == 1).all()])
    outputs_cols = list(df_exp.columns[(df_exp == 2).all()])
    
    input_df = df.loc[:, input_cols]

    # Select columns of object data type (categorical columns)
    categorical_cols = input_df.select_dtypes('object').columns.tolist()

    # Identify discrete numeric columns as categorical if they have fewer unique values than the threshold
    discrete_numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if len(input_df[col].unique()) < unique_value_threshold]
    categorical_cols.extend(discrete_numeric_cols)

    # Exclude the identified discrete numeric columns from numeric_cols
    numeric_cols = [col for col in input_df.select_dtypes(include=np.number).columns if col not in discrete_numeric_cols]
    
    return input_cols, outputs_cols, categorical_cols, numeric_cols

def save_dictionary(dict_file, file_type=".json"):
    # get name of the project as part of file name
    project_name = list(dict_file.keys())[0]
    
    if file_type == ".json":
        # Convert Python to JSON  
        json_data = json.dumps(dict_file, indent = 4)

        # write the JSON string to a file
        with open(project_name + '.json', 'w') as f:
            f.write(json_data)
            
    elif file_type == ".txt":
        # Path to save the file
        file_path = project_name + '.txt'
        
        # Convert the dictionary to a string representation
        dict_str = str(dict_file)

        # Save the string representation to a text file
        with open(file_path, 'w') as file:
            file.write(dict_str)
            
    # Return the name of the file
    filename = project_name + file_type
    print(filename)
    return filename

def load_image(image_path):
    try:
        # Open the image file
        img = Image.open(image_path)
            
        # Display basic information about the image
        print("Image Format:", img.format)
        print("Image Mode:", img.mode)
        print("Image Size:", img.size)
        
        # Show the image
        #img.show()
        
        # Display the image directly within the notebook
        display(img)
        
    except FileNotFoundError:
        print("File not found. Please check the file path.")

def merge_images(image_paths, output_path):
    images = [Image.open(path) for path in image_paths]
    
    # Determine common height if not provided
    common_height = max(img.size[1] for img in images)

    # Resize images to the common height
    resized_images = [img.resize((int(img.size[0] * common_height / img.size[1]), common_height)) for img in images]

    # Get total width of the merged image
    total_width = sum(img.size[0] for img in resized_images)

    # Create a new blank image with dimensions large enough to accommodate all images
    merged_image = Image.new("RGBA", (total_width, common_height))

    # Paste each resized image onto the blank image
    x_offset = 0
    for img in resized_images:
        merged_image.paste(img, (x_offset, 0))
        x_offset += img.size[0]

    # Save the merged image
    merged_image.save(output_path)
    
    merged_image.show()

def merge_images_grid(image_paths, output_path, columns=3):
    images = [Image.open(path) for path in image_paths]
    
    # Determine common height if not provided
    common_height = max(img.size[1] for img in images)

    # Resize images to the common height
    resized_images = [img.resize((int(img.size[0] * common_height / img.size[1]), common_height)) for img in images]
    
    # Calculate number of rows needed
    rows = (len(resized_images) + columns - 1) // columns

    # Get total dimensions of the merged image
    total_width = max(img.size[0] for img in resized_images) * columns
    total_height = common_height * rows

    # Create a new blank image with dimensions large enough to accommodate all images
    merged_image = Image.new("RGBA", (total_width, total_height))

    # Paste each resized image onto the blank image
    x_offset = 0
    y_offset = 0
    for img in resized_images:
        merged_image.paste(img, (x_offset, y_offset))
        x_offset += img.size[0]
        if x_offset >= total_width:
            x_offset = 0
            y_offset += common_height

    # Save the merged image
    merged_image.save(output_path)
    
    merged_image.show()


def check_same_feature_set(df1, df2):
    # Get the column names of both DataFrames
    features_df1 = set(df1.columns)
    features_df2 = set(df2.columns)
    
    # Check if the sets of column names are equal
    if features_df1 == features_df2:
        print("Both DataFrames have the same feature set.")
        return True
    else:
        print("The feature sets of the DataFrames are not the same.")
        different_columns_df1 = features_df1 - features_df2
        different_columns_df2 = features_df2 - features_df1
        
        if different_columns_df1:
            print("Columns in df1 but not in df2:", different_columns_df1)
        if different_columns_df2:
            print("Columns in df2 but not in df1:", different_columns_df2)
        
        return False
